# from flask import Flask, render_template, request, send_file
# import os
# from PyPDF2 import PdfFileWriter, PdfFileReader
# import io
# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import letter
# from reportlab.pdfbase.ttfonts import TTFont
# from reportlab.pdfbase import pdfmetrics
# from reportlab.pdfbase.pdfmetrics import registerFontFamily
# import fitz
# import openpyxl
# from flask import send_from_directory
# from os import path

from flask import Flask, render_template, request
from werkzeug.utils import secure_filename
import os
from PyPDF2 import PdfFileWriter, PdfFileReader
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
import fitz
import openpyxl
# from flask_cors import CORS
from flask import *
from collections import Counter
import pandas as pd
import numpy as np
import math
import gspread


# Указываем путь к JSON
gc = gspread.service_account(filename='/home/avangard3141/mysite/keys/mypython-374908-4480952f882c.json')


app = Flask(__name__)




UPLOAD_FOLDER = '/home/avangard3141/mysite/static/files/'
ALLOWED_EXTENSIONS = set(['pdf', 'jpg', 'jpeg', 'gif'])
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# app.config['C:\Users\79858\Documents\Flask-App-F\UPLOAD_FOLDER']

menu = [
    {"name": "Главная", "url": "index"},
    # {"name": "Не главная", "url": "first-app"},
    {"name": "Конвертация", "url": "contact"},
]

@app.route("/index")
@app.route("/")
def index():
    # return render_template('index.html', menu=menu)
    return render_template('contact.html', menu=menu)


@app.route("/about")
def about():
    return render_template('about.html', title="О сайте", menu=menu)

@app.route("/contact", methods=["POST", "GET"])
def contact():
    if request.method == 'POST':
        print(request.form)
    return render_template('contact.html', title="Конвертация", menu=menu)

@app.route("/base")
def base():
    return render_template('base.html', title="Base", menu=menu)

@app.route('/upload')
def upload_file():
   return render_template('upload.html')


@app.route('/updategs')
def updategs():
   return render_template('index.html')


@app.route('/uploader', methods = ['GET', 'POST'])
def uploader_file():
    if request.method == 'POST':

        folder_files_saving = "/home/avangard3141/mysite/static/files/"
        folder_files_data = "/home/avangard3141/mysite/static/data/"
        folder_white_page = "/home/avangard3141/mysite/static/data/"

        f = request.files['file']
        f.save(os.path.join(app.config['UPLOAD_FOLDER'], f.filename))
        filename_pdf = f.filename

        f1 = request.files['file1']
        f1.save(os.path.join(app.config['UPLOAD_FOLDER'], f1.filename))
        filename_excel= f1.filename

        # Если в первое окно загрузили не тот формат
        if filename_pdf[len(filename_pdf)-3:len(filename_pdf)] == "pdf":
            sheet_export_ya = pd.read_excel(folder_files_saving + filename_excel, header=0, skiprows=1)
            pdf = fitz.open(folder_files_saving + filename_pdf)
            existing_pdf = PdfFileReader(open(folder_files_saving + filename_pdf, "rb"))
        else:
            sheet_export_ya = pd.read_excel(folder_files_saving + filename_pdf, header=0, skiprows=1)
            pdf = fitz.open(folder_files_saving + filename_excel)
            existing_pdf = PdfFileReader(open(folder_files_saving + filename_excel, "rb"))

        white_page = PdfFileReader(open(folder_white_page + "white_page.pdf", "rb"))
        # Сортирую таблицу экспорта по SKU
        sheet_export_ya.sort_values(by=['Ваш SKU'], inplace=True)

        sheet_sku_data_base = openpyxl.open(folder_files_data + "sku-data-base.xlsx").active
        output = PdfFileWriter()

        # Массив, в который запишу номера страниц из PDF, которые есть в таблице экспорта
        searched_pages_on_pdf = []



        # Цикл в котором пробегаю по столбцу с номерами заказов из таблицы экспорта
        for i in range(0, len(sheet_export_ya)):
            # Переменная, хранящая номер заказа из таблицы экспорта
            # search_sku_in_sheet_ya = sheet_export_ya[i][1].value
            search_sku_in_sheet_ya = str(sheet_export_ya.iloc[i]['Ваш номер заказа'])
            # Цикл в котором пробегаю по каждой странице в pdf и ищу на какой странице этот номер заказа
            for current_page in range(len(pdf)):
                page = pdf.load_page(current_page)
                # Если я нашел страницу на которой этот номер заказа
                if page.search_for(search_sku_in_sheet_ya):
                    # print('\n' + '%s найдено на %i странице' % (search_sku_in_sheet_ya, current_page + 1))
                    searched_pages_on_pdf.append(current_page)

                    print("Найдены страницы: ", searched_pages_on_pdf)

                    # Переменная счетчик, если в базе не будет найдет этот артикул, то подставь оригинальный
                    counter_my = 0

                    # Беру sku этого заказа и ищу его в базе транслейта
                    for j in range(2, sheet_sku_data_base.max_row+1):

                        # Если я нашел этот sku в базе, то пишу его на странице pdf
                        if (sheet_export_ya.iloc[i]['Ваш SKU']) == (sheet_sku_data_base[j][1].value):

                            packet1 = io.BytesIO()
                            can = canvas.Canvas(packet1, pagesize=letter)
                            pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                            can.setFont('Roboto', 6)
                            can.rotate(90)

                            can.drawString(
                                120,
                                -338,
                                str((sheet_sku_data_base[j][2].value)) +
                                ' (' +
                                str(int(sheet_export_ya.iloc[i]['Количество'])) +
                                ' шт.)'
                            )

                            print(
                                str((sheet_sku_data_base[j][2].value)) +
                                ' (' +
                                str(int(sheet_export_ya.iloc[i]['Количество'])) +
                                ' шт.)'
                            )


                            can.save()

                            #move to the beginning of the StringIO buffer
                            packet1.seek(0)
                            new_pdf1 = PdfFileReader(packet1)

                            # add the "watermark" (which is the new pdf) on the existing page
                            page = existing_pdf.getPage(current_page)
                            page.mergePage(new_pdf1.getPage(0))
                            output.addPage(page)
                            counter_my = counter_my + 1

                        if ((sheet_sku_data_base.max_row+1) == (j + 1)) and (counter_my == 0):
                            print("я тутутутуту ")
                            packet1 = io.BytesIO()
                            can = canvas.Canvas(packet1, pagesize=letter)
                            pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                            can.setFont('Roboto', 6)
                            can.rotate(90)

                            can.drawString(
                                120,
                                -338,
                                str(sheet_export_ya.iloc[i]['Ваш SKU']) +
                                ' (' +
                                str(int(sheet_export_ya.iloc[i]['Количество'])) +
                                ' шт.)'
                            )

                            print(
                                str((sheet_sku_data_base[j][2].value)) +
                                ' (' +
                                str(int(sheet_export_ya.iloc[i]['Количество'])) +
                                ' шт.)'
                            )

                            can.save()

                            #move to the beginning of the StringIO buffer
                            packet1.seek(0)
                            new_pdf1 = PdfFileReader(packet1)

                            # add the "watermark" (which is the new pdf) on the existing page
                            page = existing_pdf.getPage(current_page)
                            page.mergePage(new_pdf1.getPage(0))
                            output.addPage(page)


        # Цикл, который добавит в конец PDF страницы, заказов которых нет в экспорте с "Нет данных"
        for current_page in range(len(pdf)):
            # Если этой страницы нет в массиве найденых таблиц
            if current_page not in searched_pages_on_pdf:

                print("Этого нет:",  current_page)

                packet3 = io.BytesIO()
                can = canvas.Canvas(packet3, pagesize=letter)
                pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                can.setFont('Roboto', 6)
                can.rotate(90)

                can.drawString(
                    120,
                    -338,
                    'Нет данных'
                )

                can.save()

                #move to the beginning of the StringIO buffer
                packet3.seek(1)
                new_pdf3 = PdfFileReader(packet3)

                page = existing_pdf.getPage(current_page)
                page.mergePage(new_pdf3.getPage(0))
                output.addPage(page)

        # Перевожу таблицу экспорта из яндекса в DataFrame для сводной таблицы
        df = sheet_export_ya
        # Сводная таблица
        df_pivot = df.pivot_table(values=['Количество'], index='Ваш SKU', aggfunc='sum', margins= True , margins_name='Sum').reset_index()
        print(df_pivot)

        # Сколько белых страниц понадобится для вывода всей таблицы
        count_white_pages = math.ceil(len(df_pivot)/35)
        print("Вот столько листов нада: ", count_white_pages)

        # Цикл, который выводит таблицу в конце PDF
        for white_pages_num in range(0, count_white_pages):

            packet4 = io.BytesIO()
            can_white_page = canvas.Canvas(packet4, pagesize=letter)

            # Размер шрифта
            font_size_white_page = 8
            # Размер строки
            line_size_white_page = font_size_white_page + 1
            # Длиина линии
            line_long = 210

            can_white_page.setFont('Roboto', font_size_white_page)
            can_white_page.rotate(90)

            # Строка заголовков
            can_white_page.line(
                line_long,  #Длина
                - 12,
                15, #Начало от левого края
                - 12,
            )
            can_white_page.drawString(
                20,
                -22,
                "Кол-во"
            )
            can_white_page.drawString(
                55,
                -22,
                "Артикул"
            )
            can_white_page.drawString(
                138,
                -22,
                "SKU"
            )
            can_white_page.line(
                line_long,  #Длина
                - 26,
                15, #Начало от левого края
                - 26,
            )

            # Переменная-счетчик для номера строки и переноса на новую страницу 36-ой строчки
            line_num = 0

            for total_pivot in range(white_pages_num*35, (white_pages_num+1)*35):

                # Выход из цыкла на последней строчке таблицы
                if total_pivot == len(df_pivot):
                    break

                # Столбец количество
                can_white_page.drawString(
                        20,
                        -33 - (line_num * line_size_white_page),
                        str(df_pivot.iat[total_pivot,1])
                    )

                # Столбец артикулы
                # Если последняя строка, то выведу итоги
                if total_pivot == len(df_pivot)-1:
                    can_white_page.drawString(
                        55,
                        -33 - (line_num * line_size_white_page),
                        "ИТОГО ТОВАРОВ"
                    )
                else:
                    for j in range(2, sheet_sku_data_base.max_row+1):
                        # Если я нашел этот sku в базе, то пишу его на странице pdf
                        if (str(df_pivot.iat[total_pivot,0])) == (sheet_sku_data_base[j][1].value):
                            can_white_page.drawString(
                                    55,
                                    -33 - (line_num  * line_size_white_page),
                                    str(sheet_sku_data_base[j][2].value)
                            )

                # Столбец артикуры изначально (SKU)
                if total_pivot < len(df_pivot)-1:
                    can_white_page.drawString(
                        138,
                        -33 - (line_num  * line_size_white_page),
                        str(df_pivot.iat[total_pivot,0])
                    )

                # Линии
                can_white_page.line(
                        line_long,  #Длина
                        - 34 - (line_num * line_size_white_page + 1),
                        15, #Начало от левого края
                        - 34 - (line_num * line_size_white_page + 1),
                        )
                # Счетчик линий, чтобы сделать перенос таблицы на новую страницу на 36-ой строчке
                line_num = line_num + 1

            can_white_page.save()

            #move to the beginning of the StringIO buffer
            packet4.seek(1)
            new_pdf4 = PdfFileReader(packet4)

            page = white_page.getPage(white_pages_num)
            page.mergePage(new_pdf4.getPage(0))
            output.addPage(page)

        # finally, write "output" to a real file

        outputStream = open("/home/avangard3141/mysite/addedindexes.pdf", "wb")
        output.write(outputStream)
        outputStream.close()
        return render_template('download.html')


@app.route('/uploadergoogle', methods = ['GET', 'POST'])
def uploadergoogle_file():
    if request.method == 'POST':

        folder_files_saving = "/home/avangard3141/mysite/static/files/"
        folder_files_data = "/home/avangard3141/mysite/static/data/"
        folder_white_page = "/home/avangard3141/mysite/static/data/"

        f = request.files['file2']
        f.save(os.path.join(app.config['UPLOAD_FOLDER'], f.filename))
        filename_pdf = f.filename

        f1 = request.files['file3']
        f1.save(os.path.join(app.config['UPLOAD_FOLDER'], f1.filename))
        filename_excel= f1.filename

        # Если в первое окно загрузили не тот формат
        if filename_pdf[len(filename_pdf)-3:len(filename_pdf)] == "pdf":
            sheet_export_ya = pd.read_excel(folder_files_saving + filename_excel, header=0, skiprows=1)
            pdf = fitz.open(folder_files_saving + filename_pdf)
            existing_pdf = PdfFileReader(open(folder_files_saving + filename_pdf, "rb"))
        else:
            sheet_export_ya = pd.read_excel(folder_files_saving + filename_pdf, header=0, skiprows=1)
            pdf = fitz.open(folder_files_saving + filename_excel)
            existing_pdf = PdfFileReader(open(folder_files_saving + filename_excel, "rb"))

        white_page = PdfFileReader(open(folder_white_page + "white_page.pdf", "rb"))
        # Сортирую таблицу экспорта по SKU
        sheet_export_ya.sort_values(by=['Ваш SKU'], inplace=True)

        # sheet_sku_data_base = openpyxl.open(folder_files_data + "sku-data-base.xlsx").active

        # Открываем таблицу
        sh = gc.open_by_url('https://docs.google.com/spreadsheets/d/1Vx9RkLzxtsncULEkd7XsHQgVrcS-dzQSdocZjLp8Uw0/edit#gid=0')
        # Лист, в который вставляем
        worksheet = sh.worksheet('SKU').get_all_records()
        df_sheet_sku_data_base = pd.DataFrame.from_dict(worksheet)

        print("\nДАТАФРЕЙМ")
        print(df_sheet_sku_data_base.head())

        print("\nПОИСК")
        print(df_sheet_sku_data_base.loc[df_sheet_sku_data_base['SKU'] == 'SC-A4-1-500as'])

        print("\n УСЛИ")
        if df_sheet_sku_data_base.loc[df_sheet_sku_data_base['SKU'] == 'SC-A4-1-500as'].empty:
            print("НИЧЕ")
        else:
            print("ЧЕТО ЕСТЬ")



        output = PdfFileWriter()

        # Массив, в который запишу номера страниц из PDF, которые есть в таблице экспорта
        searched_pages_on_pdf = []

        # Цикл в котором пробегаю по столбцу с номерами заказов из таблицы экспорта
        for i in range(0, len(sheet_export_ya)):
            # Переменная, хранящая номер заказа из таблицы экспорта
            # search_sku_in_sheet_ya = sheet_export_ya[i][1].value
            search_sku_in_sheet_ya = str(sheet_export_ya.iloc[i]['Ваш номер заказа'])
            # Цикл в котором пробегаю по каждой странице в pdf и ищу на какой странице этот номер заказа
            for current_page in range(len(pdf)):
                page = pdf.load_page(current_page)
                # Если я нашел страницу на которой этот номер заказа
                if page.search_for(search_sku_in_sheet_ya):
                    # print('\n' + '%s найдено на %i странице' % (search_sku_in_sheet_ya, current_page + 1))
                    searched_pages_on_pdf.append(current_page)

                    # print("Найдены страницы: ", searched_pages_on_pdf)

                    # Переменная счетчик, если в базе не будет найдет этот артикул, то подставь оригинальный
                    counter_my = 0

                    # number of rows in dataframe
                    # num_rows_df_sheet_sku_data_base = df.shape[0]

                    # Беру sku этого заказа и ищу его в базе транслейта
                    # for j in range(2, num_rows_df_sheet_sku_data_base+1):

                    find = df_sheet_sku_data_base.loc[df_sheet_sku_data_base['SKU'] == sheet_export_ya.iloc[i]['Ваш SKU']]

                    # Если я нашел этот sku в базе, то пишу его на странице pdf
                    if not find.empty:
                        print("НАЙДЕТООООО")
                        print(find.values[0][2])

                        packet1 = io.BytesIO()
                        can = canvas.Canvas(packet1, pagesize=letter)
                        pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                        can.setFont('Roboto', 6)
                        can.rotate(90)

                        can.drawString(
                            120,
                            -338,
                            str((find.values[0][2])) +
                            ' (' +
                            str(int(sheet_export_ya.iloc[i]['Количество'])) +
                            ' шт.)'
                        )

                        print(
                            str((find.values[0][2])) +
                            ' (' +
                            str(int(sheet_export_ya.iloc[i]['Количество'])) +
                            ' шт.)'
                        )


                        can.save()

                        #move to the beginning of the StringIO buffer
                        packet1.seek(0)
                        new_pdf1 = PdfFileReader(packet1)

                        # add the "watermark" (which is the new pdf) on the existing page
                        page = existing_pdf.getPage(current_page)
                        page.mergePage(new_pdf1.getPage(0))
                        output.addPage(page)
                        counter_my = counter_my + 1

                    else:
                        print("я тутутутуту ")
                        packet1 = io.BytesIO()
                        can = canvas.Canvas(packet1, pagesize=letter)
                        pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                        can.setFont('Roboto', 6)
                        can.rotate(90)

                        can.drawString(
                            120,
                            -338,
                            str(sheet_export_ya.iloc[i]['Ваш SKU']) +
                            ' (' +
                            str(int(sheet_export_ya.iloc[i]['Количество'])) +
                            ' шт.)'
                        )

                        # print(
                        #     str((sheet_sku_data_base[j][2].value)) +
                        #     ' (' +
                        #     str(int(sheet_export_ya.iloc[i]['Количество'])) +
                        #     ' шт.)'
                        # )

                        can.save()

                        #move to the beginning of the StringIO buffer
                        packet1.seek(0)
                        new_pdf1 = PdfFileReader(packet1)

                        # add the "watermark" (which is the new pdf) on the existing page
                        page = existing_pdf.getPage(current_page)
                        page.mergePage(new_pdf1.getPage(0))
                        output.addPage(page)


        # Цикл, который добавит в конец PDF страницы, заказов которых нет в экспорте с "Нет данных"
        for current_page in range(len(pdf)):
            # Если этой страницы нет в массиве найденых таблиц
            if current_page not in searched_pages_on_pdf:

                print("Этого нет:",  current_page)

                packet3 = io.BytesIO()
                can = canvas.Canvas(packet3, pagesize=letter)
                pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                can.setFont('Roboto', 6)
                can.rotate(90)

                can.drawString(
                    120,
                    -338,
                    'Нет данных'
                )

                can.save()

                #move to the beginning of the StringIO buffer
                packet3.seek(1)
                new_pdf3 = PdfFileReader(packet3)

                page = existing_pdf.getPage(current_page)
                page.mergePage(new_pdf3.getPage(0))
                output.addPage(page)

        # Перевожу таблицу экспорта из яндекса в DataFrame для сводной таблицы
        df = sheet_export_ya
        # Сводная таблица
        df_pivot = df.pivot_table(values=['Количество'], index='Ваш SKU', aggfunc='sum', margins= True , margins_name='Sum').reset_index()
        print(df_pivot)

        # Сколько белых страниц понадобится для вывода всей таблицы
        count_white_pages = math.ceil(len(df_pivot)/39)
        print("Вот столько листов нада: ", count_white_pages)

        # Цикл, который выводит таблицу в конце PDF
        for white_pages_num in range(0, count_white_pages):

            packet4 = io.BytesIO()
            can_white_page = canvas.Canvas(packet4, pagesize=letter)

            # Размер шрифта
            font_size_white_page = 6
            # Размер строки
            line_size_white_page = font_size_white_page + 1
            # Длиина линии
            line_long = 210

            can_white_page.setFont('Roboto', font_size_white_page)
            can_white_page.rotate(90)

            # Строка заголовков
            can_white_page.line(
                line_long,  #Длина
                - 12,
                15, #Начало от левого края
                - 12,
            )
            can_white_page.drawString(
                17,
                -22,
                "К-во"
            )
            can_white_page.drawString(
                37,
                -22,
                "Артикул"
            )
            can_white_page.drawString(
                128,
                -22,
                "SKU"
            )
            can_white_page.line(
                line_long,  #Длина
                - 26,
                15, #Начало от левого края
                - 26,
            )

            # Переменная-счетчик для номера строки и переноса на новую страницу 36-ой строчки
            line_num = 0

            for total_pivot in range(white_pages_num*39, (white_pages_num+1)*39):

                # Выход из цыкла на последней строчке таблицы
                if total_pivot == len(df_pivot):
                    break

                # Столбец количество
                can_white_page.drawString(
                        17,
                        -33 - (line_num * line_size_white_page),
                        str(df_pivot.iat[total_pivot,1])
                    )

                # Столбец артикулы
                # Если последняя строка, то выведу итоги
                if total_pivot == len(df_pivot)-1:
                    can_white_page.drawString(
                        37,
                        -33 - (line_num * line_size_white_page),
                        "ИТОГО ТОВАРОВ"
                    )
                    print(str(df_pivot.iat[total_pivot,0]))
                else:
                    find = df_sheet_sku_data_base.loc[df_sheet_sku_data_base['SKU'] == str(df_pivot.iat[total_pivot,0])]
                    # for j in range(2, sheet_sku_data_base.max_row+1):
                    #     # Если я нашел этот sku в базе, то пишу его на странице pdf
                    #     if (str(df_pivot.iat[total_pivot,0])) == (sheet_sku_data_base[j][1].value):
                    #         can_white_page.drawString(
                    #                 55,
                    #                 -33 - (line_num  * line_size_white_page),
                    #                 str(sheet_sku_data_base[j][2].value)
                    #         )
                    if not find.empty:
                        print('else')
                        print(find.values[0])

                        if len(find.values[0][2]) > 27:
                            can_white_page.drawString(
                                        37,
                                        -33 - (line_num  * line_size_white_page),
                                        find.values[0][2][:27]+'...'
                                )
                        else:
                            can_white_page.drawString(
                                        37,
                                        -33 - (line_num  * line_size_white_page),
                                        find.values[0][2]
                                )
                    else:
                        can_white_page.drawString(
                                    37,
                                    -33 - (line_num  * line_size_white_page),
                                    '- - - НЕ НАЙДЕНО - - -'
                            )
                    # can_white_page.drawString(
                    #                 55,
                    #                 -33 - (line_num  * line_size_white_page),
                    #                 find.values[0][2]
                    #         )

                # Столбец артикуры изначально (SKU)
                if total_pivot < len(df_pivot)-1:
                    if len(str(df_pivot.iat[total_pivot,0])) > 22:
                        can_white_page.drawString(
                            128,
                            -33 - (line_num  * line_size_white_page),
                            str(df_pivot.iat[total_pivot,0])[:23]+'...'
                        )
                    else:
                        can_white_page.drawString(
                            128,
                            -33 - (line_num  * line_size_white_page),
                            str(df_pivot.iat[total_pivot,0])
                        )

                # Линии
                can_white_page.line(
                        line_long,  #Длина
                        - 34 - (line_num * line_size_white_page + 1),
                        15, #Начало от левого края
                        - 34 - (line_num * line_size_white_page + 1),
                        )
                # Счетчик линий, чтобы сделать перенос таблицы на новую страницу на 36-ой строчке
                line_num = line_num + 1

            can_white_page.save()

            #move to the beginning of the StringIO buffer
            packet4.seek(1)
            new_pdf4 = PdfFileReader(packet4)

            page = white_page.getPage(white_pages_num)
            page.mergePage(new_pdf4.getPage(0))
            output.addPage(page)


        # finally, write "output" to a real file

        outputStream = open("/home/avangard3141/mysite/addedindexes.pdf", "wb")
        output.write(outputStream)
        outputStream.close()
        return render_template('download.html')



# @app.route('/uploader', methods = ['GET', 'POST'])
# def uploader_file():
#     if request.method == 'POST':

#         folder_files_saving = "/home/avangard3141/mysite/static/files/"
#         folder_files_data = "/home/avangard3141/mysite/static/data/"

#         f = request.files['file']
#         f.save(os.path.join(app.config['UPLOAD_FOLDER'], f.filename))
#         filename_pdf = f.filename

#         f1 = request.files['file1']
#         f1.save(os.path.join(app.config['UPLOAD_FOLDER'], f1.filename))
#         filename_excel= f1.filename

#         if filename_pdf[len(filename_pdf)-3:len(filename_pdf)] == "pdf":
#             # sheet_export_ya = openpyxl.open(folder_files_saving + filename_excel).active
#             sheet_export_ya = pd.read_excel(folder_files_saving + filename_excel, header=0, skiprows=1)
#             sheet_export_ya.sort_values(by=['Ваш SKU'], inplace=True)
#             # print(df)
#             # Имя файла в котором ищу номера ярлыков
#             pdf = fitz.open(folder_files_saving + filename_pdf)
#             # read your existing PDF
#             existing_pdf = PdfFileReader(open(folder_files_saving + filename_pdf, "rb"))
#         else:
#             # sheet_export_ya = openpyxl.open(folder_files_saving + filename_pdf).active
#             sheet_export_ya = pd.read_excel(folder_files_saving + filename_pdf, header=0, skiprows=1)
#             sheet_export_ya.sort_values(by=['Ваш SKU'], inplace=True)
#             # Имя файла в котором ищу номера ярлыков
#             pdf = fitz.open(folder_files_saving + filename_excel)
#             # read your existing PDF
#             existing_pdf = PdfFileReader(open(folder_files_saving + filename_excel, "rb"))

#         sheet_sku_data_base = openpyxl.open(folder_files_data + "sku-data-base.xlsx").active
#         output = PdfFileWriter()

#         # print("ТАК ТАК ТК: " , (sheet_export_ya.max_row))

#         musical_notes = []
#         for i in range(0, len(sheet_export_ya)):
#             musical_notes.append(int(sheet_export_ya.iloc[i]['Ваш номер заказа']))

#         c = Counter(musical_notes)
#         print("И ЧО??????????? - ", c)

#         searched_pages_on_pdf = []

#         # Цикл в котором пробегаю по столбцу с номерами заказов из таблицы экспорта
#         for i in range(0, len(sheet_export_ya)):
#             # Переменная, хранящая номер заказа из таблицы экспорта
#             # search_sku_in_sheet_ya = sheet_export_ya[i][1].value
#             search_sku_in_sheet_ya = str(sheet_export_ya.iloc[i]['Ваш номер заказа'])
#             # Цикл в котором пробегаю по каждой странице в pdf и ищу на какой странице этот номер заказа
#             for current_page in range(len(pdf)):
#                 page = pdf.load_page(current_page)
#                 # Если я нашел страницу на которой этот номер заказа
#                 if page.search_for(search_sku_in_sheet_ya):
#                     print('\n' + '%s найдено на %i странице' % (search_sku_in_sheet_ya, current_page + 1))
#                     searched_pages_on_pdf.append(current_page)
#                     print("Найдены страницы: ", searched_pages_on_pdf)
#                     # Беру sku этого заказа и ищу его в базе транслейта
#                     for j in range(2, sheet_sku_data_base.max_row+1):
#                         # Если я нашел этот sku в базе, то пишу его на странице pdf
#                         if (sheet_export_ya.iloc[i]['Ваш SKU']) == (sheet_sku_data_base[j][1].value):

#                             # print(
#                             #     "Это 1: ", (sheet_export_ya[i][3].value),"\n",
#                             #     "Это 2: ", (sheet_sku_data_base[j][1].value),"\n",
#                             #     "Это 3: ", str((sheet_sku_data_base[j][2].value)),"\n",
#                             # )

#                             packet1 = io.BytesIO()
#                             can = canvas.Canvas(packet1, pagesize=letter)
#                             pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
#                             can.setFont('Roboto', 6)
#                             can.rotate(90)
#                             print(
#                                         "Цикл: " +
#                                         str(int(i))
#                                         # ", " +
#                                         # (sheet_export_ya[i][1].value) +
#                                         # ", " +
#                                         # (sheet_export_ya[i+1][1].value)
#                                     )

#                             can.drawString(
#                                 120,
#                                 -338,
#                                 str((sheet_sku_data_base[j][2].value)) +
#                                 ' (' +
#                                 str(int(sheet_export_ya.iloc[i]['Количество'])) +
#                                 'pcs)'
#                             )
#                             print(
#                                 str((sheet_sku_data_base[j][2].value)) +
#                                 ' (' +
#                                 str(int(sheet_export_ya.iloc[i]['Количество'])) +
#                                 'pcs)'
#                             )

#                             can.save()

#                             #move to the beginning of the StringIO buffer
#                             packet1.seek(0)
#                             new_pdf1 = PdfFileReader(packet1)

#                             # add the "watermark" (which is the new pdf) on the existing page
#                             page = existing_pdf.getPage(current_page)
#                             page.mergePage(new_pdf1.getPage(0))
#                             output.addPage(page)
#                 # else:
#                 #     page = existing_pdf.getPage(current_page)
#                 #     page.mergePage(new_pdf1.getPage(0))
#                 #     output.addPage(page)


#         for current_page in range(len(pdf)):
#             # for searched_pages in searched_pages_on_pdf:
#             if current_page not in searched_pages_on_pdf:
#                 print("Этого нет:",  current_page)

#                 packet3 = io.BytesIO()
#                 can = canvas.Canvas(packet3, pagesize=letter)
#                 pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
#                 can.setFont('Roboto', 6)
#                 can.rotate(90)

#                 can.drawString(
#                     120,
#                     -338,
#                     'Нет данных'
#                 )

#                 can.save()

#                 #move to the beginning of the StringIO buffer
#                 packet3.seek(1)
#                 new_pdf3 = PdfFileReader(packet3)

#                 page = existing_pdf.getPage(current_page)
#                 page.mergePage(new_pdf3.getPage(0))
#                 output.addPage(page)


         # finally, write "output" to a real file
        # outputStream = open("/home/avangard3141/mysite/addedindexes.pdf", "wb")
        # output.write(outputStream)
        # outputStream.close()
        # return render_template('download.html')

# @app.route('/uploader', methods = ['GET', 'POST'])
# def uploader_file():
#     if request.method == 'POST':

#         folder_files_saving = "/home/avangard3141/mysite/static/files/"
#         folder_files_data = "/home/avangard3141/mysite/static/data/"


#         f = request.files['file']
#         f.save(os.path.join(app.config['UPLOAD_FOLDER'], f.filename))
#         filename_pdf = f.filename

#         f1 = request.files['file1']
#         f1.save(os.path.join(app.config['UPLOAD_FOLDER'], f1.filename))
#         filename_excel= f1.filename

#         if filename_pdf[len(filename_pdf)-3:len(filename_pdf)] == "pdf":
#             sheet_export_ya = openpyxl.open(folder_files_saving + filename_excel).active
#             # Имя файла в котором ищу номера ярлыков
#             pdf = fitz.open(folder_files_saving + filename_pdf)
#             # read your existing PDF
#             existing_pdf = PdfFileReader(open(folder_files_saving + filename_pdf, "rb"))
#         else:
#             sheet_export_ya = openpyxl.open(folder_files_saving + filename_pdf).active
#             # Имя файла в котором ищу номера ярлыков
#             pdf = fitz.open(folder_files_saving + filename_excel)
#             # read your existing PDF
#             existing_pdf = PdfFileReader(open(folder_files_saving + filename_excel, "rb"))

#         sheet_sku_data_base = openpyxl.open(folder_files_data + "sku-data-base.xlsx").active
#         output = PdfFileWriter()

#         # Массив в котором ищу на какой странице этот ярлык
#         for i in range(3, sheet_export_ya.max_row+1):
#             search_sku_in_sheet_ya = sheet_export_ya[i][1].value
#             for current_page in range(len(pdf)):
#                 page = pdf.load_page(current_page)

#                 if page.search_for(search_sku_in_sheet_ya):
#                     print('%s найдено на %i странице' % (search_sku_in_sheet_ya, current_page + 1))

#                     for j in range(2, sheet_sku_data_base.max_row+1):
#                         if (sheet_export_ya[i][3].value) == (sheet_sku_data_base[j][1].value):
#                             print(
#                                 "Это 1: ", (sheet_export_ya[i][3].value),"\n",
#                                 "Это 2: ", (sheet_sku_data_base[j][1].value),"\n",
#                                 "Это 3: ", str((sheet_sku_data_base[j][2].value)),"\n",
#                             )

#                             packet1 = io.BytesIO()
#                             can = canvas.Canvas(packet1, pagesize=letter)
#                             pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
#                             can.setFont('Roboto', 6)

#                             can.rotate(90)
#                             can.drawString(120, -338, str((sheet_sku_data_base[j][2].value)) + ' (' + str(int(sheet_export_ya[i][5].value)) + 'pcs)')
#                             can.save()

#                             #move to the beginning of the StringIO buffer
#                             packet1.seek(0)
#                             new_pdf1 = PdfFileReader(packet1)

#                             # add the "watermark" (which is the new pdf) on the existing page
#                             page = existing_pdf.getPage(current_page)
#                             page.mergePage(new_pdf1.getPage(0))
#                             output.addPage(page)

#          # finally, write "output" to a real file
#         outputStream = open("/home/avangard3141/mysite/addedindexes.pdf", "wb")
#         output.write(outputStream)
#         outputStream.close()
#         return render_template('download.html')


@app.route('/download')
def download():
    filename = 'addedindexes.pdf'
    return send_file(filename,as_attachment=True)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)