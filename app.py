import os
from flask import Flask, render_template, request, send_file
from PyPDF2 import PdfFileWriter, PdfFileReader
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import fitz
import openpyxl
from flask_cors import CORS
import pandas as pd
import math
import gspread


# Указываем путь к JSON
gc = gspread.service_account(filename='keys/mypython-374908-4480952f882c.json')
# gc = gspread.service_account(filename='/home/avangard3141/mysite/keys/mypython-374908-4480952f882c.json')


app = Flask(__name__)
cors = CORS(app)

# Для pythonanywhere
# UPLOAD_FOLDER = "/home/avangard3141/mysite/static/files/"
# FOLDER_FILES_SAVE = "/home/avangard3141/mysite/static/files/"
# FOLDER_WHITE_PAGE = "/home/avangard3141/mysite/static/data/"
# FOLDER_FILES_DATA = "/home/avangard3141/mysite/static/data/"
# FOLDER_OUTPUT_STREAM = "/home/avangard3141/mysite/addedindexes.pdf"

# Для локалки
UPLOAD_FOLDER = "C:/Users/79858/Documents/1-git/Ya-Market-Web-App/static/files/"
FOLDER_FILES_SAVE = "C:/Users/79858/Documents/1-git/Ya-Market-Web-App/static/files/"
FOLDER_WHITE_PAGE = "C:/Users/79858/Documents/1-git/Ya-Market-Web-App/static/data/"
FOLDER_FILES_DATA = "C:/Users/79858/Documents/1-git/Ya-Market-Web-App/static/data/"
FOLDER_OUTPUT_STREAM = "addedindexes.pdf"

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['FOLDER_FILES_SAVE'] = FOLDER_FILES_SAVE
app.config['FOLDER_WHITE_PAGE'] = FOLDER_WHITE_PAGE
app.config['FOLDER_FILES_DATA'] = FOLDER_FILES_DATA
app.config['FOLDER_OUTPUT_STREAM'] = FOLDER_OUTPUT_STREAM


@app.route("/")
def index():
    return render_template('index.html')


@app.route('/uploader', methods = ['GET', 'POST'])
def uploader_file():
    if request.method == 'POST':

        f = request.files['file']
        f.save(os.path.join(app.config['UPLOAD_FOLDER'], f.filename))
        filename_pdf = f.filename

        f1 = request.files['file1']
        f1.save(os.path.join(app.config['UPLOAD_FOLDER'], f1.filename))
        filename_excel= f1.filename

        # Если в первое окно загрузили не тот формат
        if filename_pdf[len(filename_pdf)-3:len(filename_pdf)] == "pdf":
            sheet_export_ya = pd.read_excel(FOLDER_FILES_SAVE + filename_excel, header=0, skiprows=1)
            pdf = fitz.open(FOLDER_FILES_SAVE + filename_pdf)
            existing_pdf = PdfFileReader(open(FOLDER_FILES_SAVE + filename_pdf, "rb"))
        else:
            sheet_export_ya = pd.read_excel(FOLDER_FILES_SAVE + filename_pdf, header=0, skiprows=1)
            pdf = fitz.open(FOLDER_FILES_SAVE + filename_excel)
            existing_pdf = PdfFileReader(open(FOLDER_FILES_SAVE + filename_excel, "rb"))

        white_page = PdfFileReader(open(FOLDER_WHITE_PAGE + "white_page.pdf", "rb"))
        
        # Сортирую таблицу экспорта по SKU
        sheet_export_ya.sort_values(by=['Ваш SKU'], inplace=True)

        sheet_sku_data_base = openpyxl.open(FOLDER_FILES_DATA + "sku-data-base.xlsx").active
        output = PdfFileWriter()

        # Массив, в который запишу номера страниц из PDF, которые есть в таблице экспорта
        searched_pages_on_pdf = []

        # Цикл в котором пробегаю по столбцу с номерами заказов из таблицы экспорта
        for i in range(0, len(sheet_export_ya)):
            # Переменная, хранящая номер заказа из таблицы экспорта
            # search_sku_in_sheet_ya = sheet_export_ya[i][1].value
            search_sku_in_sheet_ya = str(sheet_export_ya.iloc[i]['Ваш номер заказа'])
            # Цикл в котором пробегаю по каждой странице в pdf и ищу на какой странице этот номер заказа
            for current_page in range(len(pdf)):
                page = pdf.load_page(current_page)
                # Если я нашел страницу на которой этот номер заказа
                if page.search_for(search_sku_in_sheet_ya):
                    # print('\n' + '%s найдено на %i странице' % (search_sku_in_sheet_ya, current_page + 1))
                    searched_pages_on_pdf.append(current_page)

                    print("Найдены страницы: ", searched_pages_on_pdf)

                    # Переменная счетчик, если в базе не будет найдет этот артикул, то подставь оригинальный
                    counter_my = 0

                    # Беру sku этого заказа и ищу его в базе транслейта
                    for j in range(2, sheet_sku_data_base.max_row+1):

                        # Если я нашел этот sku в базе, то пишу его на странице pdf
                        if (sheet_export_ya.iloc[i]['Ваш SKU']) == (sheet_sku_data_base[j][1].value):

                            packet1 = io.BytesIO()
                            can = canvas.Canvas(packet1, pagesize=letter)
                            pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                            can.setFont('Roboto', 6)
                            can.rotate(90)

                            can.drawString(
                                120,
                                -338,
                                str((sheet_sku_data_base[j][2].value)) +
                                ' (' +
                                str(int(sheet_export_ya.iloc[i]['Количество'])) +
                                ' шт.)'
                            )

                            print(
                                str((sheet_sku_data_base[j][2].value)) +
                                ' (' +
                                str(int(sheet_export_ya.iloc[i]['Количество'])) +
                                ' шт.)'
                            )


                            can.save()

                            #move to the beginning of the StringIO buffer
                            packet1.seek(0)
                            new_pdf1 = PdfFileReader(packet1)

                            # add the "watermark" (which is the new pdf) on the existing page
                            page = existing_pdf.getPage(current_page)
                            page.mergePage(new_pdf1.getPage(0))
                            output.addPage(page)
                            counter_my = counter_my + 1

                        if ((sheet_sku_data_base.max_row+1) == (j + 1)) and (counter_my == 0):
                            print("я тутутутуту ")
                            packet1 = io.BytesIO()
                            can = canvas.Canvas(packet1, pagesize=letter)
                            pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                            can.setFont('Roboto', 6)
                            can.rotate(90)

                            can.drawString(
                                120,
                                -338,
                                str(sheet_export_ya.iloc[i]['Ваш SKU']) +
                                ' (' +
                                str(int(sheet_export_ya.iloc[i]['Количество'])) +
                                ' шт.)'
                            )

                            print(
                                str((sheet_sku_data_base[j][2].value)) +
                                ' (' +
                                str(int(sheet_export_ya.iloc[i]['Количество'])) +
                                ' шт.)'
                            )

                            can.save()

                            #move to the beginning of the StringIO buffer
                            packet1.seek(0)
                            new_pdf1 = PdfFileReader(packet1)

                            # add the "watermark" (which is the new pdf) on the existing page
                            page = existing_pdf.getPage(current_page)
                            page.mergePage(new_pdf1.getPage(0))
                            output.addPage(page)


        # Цикл, который добавит в конец PDF страницы, заказов которых нет в экспорте с "Нет данных"
        for current_page in range(len(pdf)):
            # Если этой страницы нет в массиве найденых таблиц
            if current_page not in searched_pages_on_pdf:

                print("Этого нет:",  current_page)

                packet3 = io.BytesIO()
                can = canvas.Canvas(packet3, pagesize=letter)
                pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                can.setFont('Roboto', 6)
                can.rotate(90)

                can.drawString(
                    120,
                    -338,
                    'Нет данных'
                )

                can.save()

                #move to the beginning of the StringIO buffer
                packet3.seek(1)
                new_pdf3 = PdfFileReader(packet3)

                page = existing_pdf.getPage(current_page)
                page.mergePage(new_pdf3.getPage(0))
                output.addPage(page)

        # Перевожу таблицу экспорта из яндекса в DataFrame для сводной таблицы
        df = sheet_export_ya
        # Сводная таблица
        df_pivot = df.pivot_table(values=['Количество'], index='Ваш SKU', aggfunc='sum', margins= True , margins_name='Sum').reset_index()
        print(df_pivot)

        # Сколько белых страниц понадобится для вывода всей таблицы
        count_white_pages = math.ceil(len(df_pivot)/35)
        print("Вот столько листов нада: ", count_white_pages)

        # Цикл, который выводит таблицу в конце PDF
        for white_pages_num in range(0, count_white_pages):

            packet4 = io.BytesIO()
            can_white_page = canvas.Canvas(packet4, pagesize=letter)

            # Размер шрифта
            font_size_white_page = 8
            # Размер строки
            line_size_white_page = font_size_white_page + 1
            # Длиина линии
            line_long = 210

            can_white_page.setFont('Roboto', font_size_white_page)
            can_white_page.rotate(90)

            # Строка заголовков
            can_white_page.line(
                line_long,  #Длина
                - 12,
                15, #Начало от левого края
                - 12,
            )
            can_white_page.drawString(
                20,
                -22,
                "Кол-во"
            )
            can_white_page.drawString(
                55,
                -22,
                "Артикул"
            )
            can_white_page.drawString(
                138,
                -22,
                "SKU"
            )
            can_white_page.line(
                line_long,  #Длина
                - 26,
                15, #Начало от левого края
                - 26,
            )

            # Переменная-счетчик для номера строки и переноса на новую страницу 36-ой строчки
            line_num = 0

            for total_pivot in range(white_pages_num*35, (white_pages_num+1)*35):

                # Выход из цыкла на последней строчке таблицы
                if total_pivot == len(df_pivot):
                    break

                # Столбец количество
                can_white_page.drawString(
                        20,
                        -33 - (line_num * line_size_white_page),
                        str(df_pivot.iat[total_pivot,1])
                    )

                # Столбец артикулы
                # Если последняя строка, то выведу итоги
                if total_pivot == len(df_pivot)-1:
                    can_white_page.drawString(
                        55,
                        -33 - (line_num * line_size_white_page),
                        "ИТОГО ТОВАРОВ"
                    )
                else:
                    for j in range(2, sheet_sku_data_base.max_row+1):
                        # Если я нашел этот sku в базе, то пишу его на странице pdf
                        if (str(df_pivot.iat[total_pivot,0])) == (sheet_sku_data_base[j][1].value):
                            can_white_page.drawString(
                                    55,
                                    -33 - (line_num  * line_size_white_page),
                                    str(sheet_sku_data_base[j][2].value)
                            )

                # Столбец артикуры изначально (SKU)
                if total_pivot < len(df_pivot)-1:
                    can_white_page.drawString(
                        138,
                        -33 - (line_num  * line_size_white_page),
                        str(df_pivot.iat[total_pivot,0])
                    )

                # Линии
                can_white_page.line(
                        line_long,  #Длина
                        - 34 - (line_num * line_size_white_page + 1),
                        15, #Начало от левого края
                        - 34 - (line_num * line_size_white_page + 1),
                        )
                # Счетчик линий, чтобы сделать перенос таблицы на новую страницу на 36-ой строчке
                line_num = line_num + 1

            can_white_page.save()

            #move to the beginning of the StringIO buffer
            packet4.seek(1)
            new_pdf4 = PdfFileReader(packet4)

            page = white_page.getPage(white_pages_num)
            page.mergePage(new_pdf4.getPage(0))
            output.addPage(page)

        # finally, write "output" to a real file

        outputStream = open(FOLDER_OUTPUT_STREAM, "wb")
        output.write(outputStream)
        outputStream.close()
        return render_template('download.html')


@app.route('/uploadergoogle', methods = ['GET', 'POST'])
def uploadergoogle_file():
    if request.method == 'POST':

        f = request.files['file2']
        f.save(os.path.join(UPLOAD_FOLDER, f.filename))
        filename_pdf = f.filename

        f1 = request.files['file3']
        f1.save(os.path.join(UPLOAD_FOLDER, f1.filename))
        filename_excel= f1.filename

        # Если в первое окно загрузили не тот формат
        if filename_pdf[len(filename_pdf)-3:len(filename_pdf)] == "pdf":
            sheet_export_ya = pd.read_excel(FOLDER_FILES_SAVE + filename_excel, header=0, skiprows=1)
            pdf = fitz.open(FOLDER_FILES_SAVE + filename_pdf)
            existing_pdf = PdfFileReader(open(FOLDER_FILES_SAVE + filename_pdf, "rb"))
        else:
            sheet_export_ya = pd.read_excel(FOLDER_FILES_SAVE + filename_pdf, header=0, skiprows=1)
            pdf = fitz.open(FOLDER_FILES_SAVE + filename_excel)
            existing_pdf = PdfFileReader(open(FOLDER_FILES_SAVE + filename_excel, "rb"))

        white_page = PdfFileReader(open(FOLDER_WHITE_PAGE + "white_page.pdf", "rb"))
        # Сортирую таблицу экспорта по SKU
        sheet_export_ya.sort_values(by=['Ваш SKU'], inplace=True)

        # Открываем таблицу
        sh = gc.open_by_url('https://docs.google.com/spreadsheets/d/1Vx9RkLzxtsncULEkd7XsHQgVrcS-dzQSdocZjLp8Uw0/edit#gid=0')
        # Лист, в который вставляем
        worksheet = sh.worksheet('SKU').get_all_records()
        df_sheet_sku_data_base = pd.DataFrame.from_dict(worksheet)

        print("\nТАБЛИЦА СО SKU ИЗ ГУГЛА")
        print(df_sheet_sku_data_base.head())

        output = PdfFileWriter()

        # Массив, в который запишу номера страниц из PDF, которые есть в таблице экспорта
        searched_pages_on_pdf = []

        # Цикл в котором пробегаю по столбцу с номерами заказов из таблицы экспорта
        for i in range(0, len(sheet_export_ya)):
            # Переменная, хранящая номер заказа из таблицы экспорта
            # search_sku_in_sheet_ya = sheet_export_ya[i][1].value
            search_sku_in_sheet_ya = str(sheet_export_ya.iloc[i]['Ваш номер заказа'])
            # Цикл в котором пробегаю по каждой странице в pdf и ищу на какой странице этот номер заказа
            for current_page in range(len(pdf)):
                page = pdf.load_page(current_page)
                # Если я нашел страницу на которой этот номер заказа
                if page.search_for(search_sku_in_sheet_ya):

                    searched_pages_on_pdf.append(current_page)

                    print("Найдены страницы: ", searched_pages_on_pdf)

                    # Переменная счетчик, если в базе не будет найдет этот артикул, то подставь оригинальный
                    counter_my = 0

                    find = df_sheet_sku_data_base.loc[df_sheet_sku_data_base['SKU'] == sheet_export_ya.iloc[i]['Ваш SKU']]

                    # Если я нашел этот sku в базе, то пишу его на странице pdf
                    if not find.empty:

                        packet1 = io.BytesIO()
                        can = canvas.Canvas(packet1, pagesize=letter)
                        pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                        can.setFont('Roboto', 6)
                        can.rotate(90)

                        can.drawString(
                            120,
                            -338,
                            str((find.values[0][2])) +
                            ' (' +
                            str(int(sheet_export_ya.iloc[i]['Количество'])) +
                            ' шт.)'
                        )

                        # print(
                        #     str((find.values[0][2])) +
                        #     ' (' +
                        #     str(int(sheet_export_ya.iloc[i]['Количество'])) +
                        #     ' шт.)'
                        # )

                        can.save()

                        #move to the beginning of the StringIO buffer
                        packet1.seek(0)
                        new_pdf1 = PdfFileReader(packet1)

                        # add the "watermark" (which is the new pdf) on the existing page
                        page = existing_pdf.getPage(current_page)
                        page.mergePage(new_pdf1.getPage(0))
                        output.addPage(page)
                        counter_my = counter_my + 1

                    else:
                        packet1 = io.BytesIO()
                        can = canvas.Canvas(packet1, pagesize=letter)
                        pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                        can.setFont('Roboto', 6)
                        can.rotate(90)

                        can.drawString(
                            120,
                            -338,
                            str(sheet_export_ya.iloc[i]['Ваш SKU']) +
                            ' (' +
                            str(int(sheet_export_ya.iloc[i]['Количество'])) +
                            ' шт.)'
                        )

                        # print(
                        #     str((sheet_sku_data_base[j][2].value)) +
                        #     ' (' +
                        #     str(int(sheet_export_ya.iloc[i]['Количество'])) +
                        #     ' шт.)'
                        # )

                        can.save()

                        #move to the beginning of the StringIO buffer
                        packet1.seek(0)
                        new_pdf1 = PdfFileReader(packet1)

                        # add the "watermark" (which is the new pdf) on the existing page
                        page = existing_pdf.getPage(current_page)
                        page.mergePage(new_pdf1.getPage(0))
                        output.addPage(page)


        # Цикл, который добавит в конец PDF страницы, заказов которых нет в экспорте с "Нет данных"
        for current_page in range(len(pdf)):
            # Если этой страницы нет в массиве найденых таблиц
            if current_page not in searched_pages_on_pdf:

                print("Этого нет:",  current_page)

                packet3 = io.BytesIO()
                can = canvas.Canvas(packet3, pagesize=letter)
                pdfmetrics.registerFont(TTFont('Roboto', 'Roboto-Medium.ttf'))
                can.setFont('Roboto', 6)
                can.rotate(90)

                can.drawString(
                    120,
                    -338,
                    'Нет данных'
                )

                can.save()

                #move to the beginning of the StringIO buffer
                packet3.seek(1)
                new_pdf3 = PdfFileReader(packet3)

                page = existing_pdf.getPage(current_page)
                page.mergePage(new_pdf3.getPage(0))
                output.addPage(page)

        # Перевожу таблицу экспорта из яндекса в DataFrame для сводной таблицы
        df = sheet_export_ya
        # Сводная таблица
        df_pivot = df.pivot_table(values=['Количество'], index='Ваш SKU', aggfunc='sum', margins= True , margins_name='Sum').reset_index()

        print("СВОДНАЯ ТАБЛИЦА ИТОГОВ ДЛЯ PDF")
        print(df_pivot.head())

        # Сколько белых страниц понадобится для вывода всей таблицы
        count_white_pages = math.ceil(len(df_pivot)/39)
        print("Вот столько листов нада: ", count_white_pages)

        # Цикл, который выводит таблицу в конце PDF
        for white_pages_num in range(0, count_white_pages):

            packet4 = io.BytesIO()
            can_white_page = canvas.Canvas(packet4, pagesize=letter)

            # Размер шрифта
            font_size_white_page = 6
            # Размер строки
            line_size_white_page = font_size_white_page + 1
            # Длиина линии
            line_long = 210

            can_white_page.setFont('Roboto', font_size_white_page)
            can_white_page.rotate(90)

            # Строка заголовков
            can_white_page.line(
                line_long,  #Длина
                - 12,
                15, #Начало от левого края
                - 12,
            )
            can_white_page.drawString(
                17,
                -22,
                "К-во"
            )
            can_white_page.drawString(
                37,
                -22,
                "Артикул"
            )
            can_white_page.drawString(
                128,
                -22,
                "SKU"
            )
            can_white_page.line(
                line_long,  #Длина
                - 26,
                15, #Начало от левого края
                - 26,
            )

            # Переменная-счетчик для номера строки и переноса на новую страницу 36-ой строчки
            line_num = 0

            for total_pivot in range(white_pages_num*39, (white_pages_num+1)*39):

                # Выход из цыкла на последней строчке таблицы
                if total_pivot == len(df_pivot):
                    break

                # Столбец количество
                can_white_page.drawString(
                        17,
                        -33 - (line_num * line_size_white_page),
                        str(df_pivot.iat[total_pivot,1])
                    )

                # Столбец артикулы
                # Если последняя строка, то выведу итоги
                if total_pivot == len(df_pivot)-1:
                    can_white_page.drawString(
                        37,
                        -33 - (line_num * line_size_white_page),
                        "ИТОГО ТОВАРОВ"
                    )
                    print(str(df_pivot.iat[total_pivot,0]))
                else:
                    find = df_sheet_sku_data_base.loc[df_sheet_sku_data_base['SKU'] == str(df_pivot.iat[total_pivot,0])]
                    if not find.empty:
                        if len(find.values[0][2]) > 27:
                            can_white_page.drawString(
                                        37, 
                                        -33 - (line_num  * line_size_white_page), 
                                        find.values[0][2][:27]+'...'   
                                )
                        else:
                            can_white_page.drawString(
                                        37, 
                                        -33 - (line_num  * line_size_white_page), 
                                        find.values[0][2]  
                                )
                    else:
                        can_white_page.drawString(
                                    37, 
                                    -33 - (line_num  * line_size_white_page), 
                                    '- - - НЕ НАЙДЕНО - - -'   
                            )

                # Столбец артикуры изначально (SKU)
                if total_pivot < len(df_pivot)-1:
                    if len(str(df_pivot.iat[total_pivot,0])) > 22:
                        can_white_page.drawString(
                            128,
                            -33 - (line_num  * line_size_white_page),
                            str(df_pivot.iat[total_pivot,0])[:23]+'...'  
                        )
                    else:
                        can_white_page.drawString(
                            128,
                            -33 - (line_num  * line_size_white_page),
                            str(df_pivot.iat[total_pivot,0]) 
                        )

                # Линии
                can_white_page.line(
                        line_long,  #Длина
                        - 34 - (line_num * line_size_white_page + 1),
                        15, #Начало от левого края
                        - 34 - (line_num * line_size_white_page + 1),
                        )
                # Счетчик линий, чтобы сделать перенос таблицы на новую страницу на 36-ой строчке
                line_num = line_num + 1

            can_white_page.save()

            #move to the beginning of the StringIO buffer
            packet4.seek(1)
            new_pdf4 = PdfFileReader(packet4)

            page = white_page.getPage(white_pages_num)
            page.mergePage(new_pdf4.getPage(0))
            output.addPage(page)

        # finally, write "output" to a real file
        outputStream = open(FOLDER_OUTPUT_STREAM, "wb")
        output.write(outputStream)
        outputStream.close()
        return render_template('download.html')

# Для локалки
@app.route('/download')
def download():
    dist_dir = 'C:/Users/79858/Documents/1-git/Ya-Market-Web-App/'
    entry = os.path.join(dist_dir, 'addedindexes.pdf')
    return send_file(entry, as_attachment=True)


# Для pythonanywhere
# @app.route('/download')
# def download():
#     filename = 'addedindexes.pdf'
#     return send_file(filename,as_attachment=True)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug = True)

    